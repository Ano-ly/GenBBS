import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.cell_range import CellRange
from src.logic.data_models import Project, CategoryHigher, CategoryLower, Element, Bar
from PySide6.QtWidgets import QFileDialog, QMessageBox
from typing import List, Dict, Any, Union
import math
from src.utils.bar_image_generator import generate_bar_image
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
import os

class ExcelExporter:
    @staticmethod
    def _collect_element_quantities(current_item: Union[Project, CategoryHigher, CategoryLower, Element], element_quantities: Dict[str, int]):
        if isinstance(current_item, Element):
            element_quantities[current_item.id] = current_item.quantity

        children_to_check = []
        if isinstance(current_item, Project):
            children_to_check.extend(current_item.categories)
        elif isinstance(current_item, CategoryHigher):
            children_to_check.extend(current_item.children)
        elif isinstance(current_item, CategoryLower):
            children_to_check.extend(current_item.elements)

        for child in children_to_check:
            ExcelExporter._collect_element_quantities(child, element_quantities)

    @staticmethod
    def _compute_adjusted_bar_properties(element_quantities: Dict[str, int], bar_dict: Dict[str, Any]) -> Dict[str, Any]:
        bar = Bar.from_dict(bar_dict)
        
        for parent_info in bar.parent_tree:
            if parent_info.get('type') == 'Element':
                parent_id = parent_info.get('id')
                break

        element_quantity = element_quantities.get(parent_id, 1)

        total_number_of_bars_adjusted = bar.number_of_bars * element_quantity
        total_length_adjusted = bar.cut_length * bar.number_of_bars * element_quantity
        total_weight_adjusted = math.ceil(bar.total_weight * element_quantity)

        if bar.bar_mark < 10:
            bar_dict["bar_mark"] = f"0{bar.bar_mark}"
        else:
            bar_dict["bar_mark"] = str(bar.bar_mark)

        bar_dict["total_no_of_bars"] = total_number_of_bars_adjusted
        bar_dict["total_length"] = total_length_adjusted
        bar_dict["total_weight"] = total_weight_adjusted
        return bar_dict

    @staticmethod
    def _collect_hierarchical_data_recursive(item: Union[Project, CategoryHigher, CategoryLower, Element, Bar],
                                            level: int,
                                            parent_path: List[str],
                                            hierarchical_rows: List[Dict[str, Any]], 
                                            element_quantities: Dict[str, int]):
        item_type = type(item).__name__
        item_name = getattr(item, 'name', '') if hasattr(item, 'name') else ''
        current_path = parent_path + [item_name]

        row_data = {
            "Type": item_type,
            "Level": level,
            "Name": item_name,
            "Path": " > ".join(current_path),
            "Quantity": 0,
            "image_path" : ""
        }
        if isinstance(item, Element):
            row_data["Quantity"] = item.quantity

        if isinstance(item, Bar):
            row_data.update(ExcelExporter._compute_adjusted_bar_properties(element_quantities, item.to_dict()))
            image_path = generate_bar_image(item.shape_code, item.lengths, item.bar_id, r"C:\Users\Amy-Jay\Desktop\programming\GenBBS\src\temp_bar_images")
            # image_path = generate_bar_image("00", {'A':1000}, 3, "fff")
            row_data["image_path"] = image_path

        hierarchical_rows.append(row_data)

        if isinstance(item, Project):
            for category in item.categories:
                ExcelExporter._collect_hierarchical_data_recursive(category, level + 1, current_path, hierarchical_rows, element_quantities)
        elif isinstance(item, CategoryHigher):
            for child in item.children:
                ExcelExporter._collect_hierarchical_data_recursive(child, level + 1, current_path, hierarchical_rows, element_quantities)
        elif isinstance(item, CategoryLower):
            for element in item.elements:
                ExcelExporter._collect_hierarchical_data_recursive(element, level + 1, current_path, hierarchical_rows, element_quantities) 
        elif isinstance(item, Element):
            btw_headers = {"Type":"Type", "Level": level + 1, "Name":"Name", "Path":"Path", "Quantity":"Quantity", "bar_mark":"Bar Mark", "diameter":"Type and Bar Size", "number_of_bars":"No. in Each", "total_no_of_bars":"Total No. of Bars", "cut_length":"Cut Length", "total_length":"Total Length", "unit_weight":"unit_weight", "total_weight":"Total Weight", "shape_code":"Shape Code", "lengths":"Lengths"}
            hierarchical_rows.append(btw_headers)
            for bar in item.bars:
                ExcelExporter._collect_hierarchical_data_recursive(bar, level + 1, current_path, hierarchical_rows, element_quantities)

    @staticmethod
    def export_project_bars_to_excel(project: Project, file_path: str):
        if not project:
            QMessageBox.warning(None, "Export Error", "No project loaded to export.")
            return

        element_quantities = {}
        ExcelExporter._collect_element_quantities(project, element_quantities)

        hierarchical_rows = []
        ExcelExporter._collect_hierarchical_data_recursive(project, 0, [], hierarchical_rows, element_quantities)

        # Determine all possible column headers
        # all_headers = set()
        # for row in hierarchical_rows:
        #     all_headers.update(row.keys())
        ordered_headers = ["Type", "Level", "Name", "Path", "Quantity", "image_path", "bar_mark", "diameter", "number_of_bars", "total_no_of_bars", "cut_length", "total_length", "unit_weight", "total_weight", "shape_code", "lengths"]
        df = pd.DataFrame(hierarchical_rows, columns=ordered_headers)

        try:
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Project Data', index=False)
                workbook = writer.book
                worksheet = writer.sheets['Project Data']

                #styles
                header_font = Font(bold=True)
                category_font = Font(bold=True, size=16, underline="single")
                element_font = Font(size=12, underline="single")
                bar_font = Font(size=10)
                project_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid") #
                normal_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

                thin_border = Border(left=Side(style='thin'),
                                    right=Side(style='thin'),
                                    top=Side(style='thin'),
                                    bottom=Side(style='thin'))

                header_row_idx = 1
                header_row = next (worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
                col_map = {name: idx for idx, name in enumerate(header_row)}
                #Headers of columns to be deleted
                target_headers = ["Type", "Level", "Path", "bar_id", "parent_tree", "Name", "unit_weight","Quantity"]
                #Current cell headers and their corresponding proper cell headers for Bar Bending Schedule
                proper_cell_headers = {"bar_mark": "Bar Mark", "bar_id": "Bar ID", "shape_code": "Shape Code", "number_of_bars":"No. in Each", "total_no_of_bars" : "Total No. of Bars" ,"total_length" : "Total Length", "cut_length" : "Cut Length", "unit_weight": "Unit Weight", "total_weight": "Total Weight", "diameter" : "Type and Bar Size", "lengths": "Lengths"}
                list_of_names = dict()
                list_of_quantities = dict()
                max_widths_of_columns = dict()

                for row in worksheet.iter_rows():
                    row_index = row[0].row
                    row_type = row[col_map["Type"]].value
                    row_level = row[col_map["Level"]].value
                    row_image = row[col_map["image_path"]].value

                    # Header row formatting
                    if row_index == 1 or row_type == "Type":
                        for cell in row:
                            cell.font = header_font
                            cell.border = thin_border                       
                    #Height formatting
                    worksheet.row_dimensions[row_index].height = 60               
                    #color, font and alignment formatting
                    center_align = Alignment(horizontal="center", vertical="center")
                    left_align = Alignment(horizontal="left", vertical="center")
                    #Store the names of sub headers
                    if row_type != "Bar" and row_type != "Type":
                        row_name = row[col_map["Name"]].value
                        list_of_names[row_index] = row_name.upper()
                        if row_type == "Element":
                            row_quantity = row[col_map["Quantity"]].value
                            list_of_quantities[row_index] = row_quantity
                    
                    fill = normal_fill
                    font = None
                    # row_name = row[col_map["Name"]].value
                    if row_type == "Project":
                        fill = project_fill
                        font = category_font
                    elif row_type == "CategoryHigher" or row_type == "CategoryLower":
                        font = category_font
                    elif row_type == "Element":
                        font = element_font
                    elif row_type == "Bar":
                        font = bar_font
                    for cell in row:
                        #Set alignment
                        column_name = next((k for k, v in col_map.items() if v == cell.column - 1), None)
                        max_widths_of_columns[column_name] = max(max_widths_of_columns.get(column_name, 0), len(str(cell.value)))
                        # print(cell.value)
                        # print(max_widths_of_columns)
                        if row_type == "Bar" or row_type == "Type":
                            cell.alignment = center_align
                        else:
                            cell.alignment = left_align
                        if row_index != header_row_idx:
                            if fill:
                                cell.fill = fill
                            if font:
                                cell.font = font
                            cell.border = thin_border    
                    if row_type == "Bar" and row_image != "":
                        image_path = row_image
                        if os.path.exists(image_path):
                            img = Image(image_path)
                            # Adjust image size if necessary
                            img.width = 250 # Example width
                            img.height = 80 # Example height
                            #Add image to relevant column
                            img_col_idx = col_map.get("image_path")
                            if img_col_idx is not None:
                                cell = row[img_col_idx]
                                anchor = cell.coordinate
                                worksheet.add_image(img, anchor)
                            # os.remove(image_path)
                    # Set outline level for grouping
                    if row_index != header_row_idx:
                        worksheet.row_dimensions[row_index].outlineLevel = int(row_level)
                        worksheet.row_dimensions[row_index].hidden = False
                        
                for col in range(worksheet.max_column, 0, -1):
                    cell_value = worksheet.cell(row=header_row_idx, column=col).value
                    #Delete irrelevant columns
                    if cell_value in target_headers:
                        worksheet.delete_cols(col)
                        continue
                for col in range(worksheet.max_column, 0, -1):
                    # print(col)
                    cell_value = worksheet.cell(row=header_row_idx, column=col).value
                    #Adjust column width
                    column = get_column_letter(col)
                    adjusted_width = max((max_widths_of_columns.get(cell_value, 0), len(proper_cell_headers.get(cell_value, cell_value)))) + 2
                    # print(f"Adej: {adjusted_width}, {cell_value}")
                    worksheet.column_dimensions[column].width = adjusted_width    
                    #Change column or header names to appropriate ones
                    if cell_value in proper_cell_headers.keys():
                        worksheet.cell(row=header_row_idx, column=col, value=proper_cell_headers[cell_value])

                #Add names of sub-headers and quantities to the first row of each sub-header group
                header_row_trimmed = next (worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
                col_map_trimmed = {name: idx for idx, name in enumerate(header_row_trimmed)}
                for row in worksheet.iter_rows():
                    row_index = row[0].row
                    if row_index != 1:
                        row_bar_mark = row[col_map_trimmed["Bar Mark"]].value
                        if row_bar_mark == "" or row_bar_mark is None:
                            qt_cell = worksheet.cell(row=row_index, column=1)
                            name_cell = worksheet.cell(row=row_index, column=2)
                            no_border = Border(left=Side(style='thin'),
                                                right=Side(style='dashed'),
                                                top=Side(style='thin'),
                                                bottom=Side(style='thin'))
                            no_border2 = Border(left=Side(style='dashed'),
                                                right=Side(style='thin'),
                                                top=Side(style='thin'),
                                                bottom=Side(style='thin'))
                            qt_cell.border = no_border
                            name_cell.border = no_border2
                            if row_index not in list_of_quantities.keys():
                                worksheet.merge_cells(start_row=row_index, start_column=2, end_row=row_index, end_column=worksheet.max_column)
                                worksheet.cell(row=row_index, column=2, value=list_of_names[row_index])
                            else:
                                worksheet.merge_cells(start_row=row_index, start_column=2, end_row=row_index, end_column=worksheet.max_column)
                                worksheet.cell(row=row_index, column=1, value=list_of_quantities[row_index])
                                right_align = Alignment(horizontal="right", vertical="center")
                                qt_cell.alignment = right_align
                                worksheet.cell(row=row_index, column=2, value=list_of_names[row_index])
                
                #Remove initial header
                # worksheet.delete_rows(1)

                
                QMessageBox.information(None, "Export Successful", f"Project data exported to {file_path}")
        except Exception as e:
            QMessageBox.critical(None, "Export Error", f"Failed to export data: {e}")

    
